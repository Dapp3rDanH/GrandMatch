# -*- coding: utf-8 -*-
"""
export_segments.py — Extract GrandparentSegments CSV from phased Excel

Reads a phased Hickman_phased.xlsx and exports contiguous grandparent
segments in the same format as the GrandparentSegments sheet used by
GrandMatch.

Usage:
    python export_segments.py                     # default: out/Hickman_phased.xlsx
    python export_segments.py path/to/file.xlsx   # specify input
"""
import sys
import os
import csv
import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUT_DIR = os.path.join(SCRIPT_DIR, "out")
DEFAULT_INPUT = os.path.join(OUT_DIR, "Hickman_phased.xlsx")
DEFAULT_OUTPUT = os.path.join(OUT_DIR, "grandparent_segments.csv")

SIBLINGS = ["Debbie", "Denny", "Donal", "Joyce", "Roger"]

KITS = {
    "Debbie": "UH1430486",
    "Denny": "BA1871856",
    "Donal": "T806693",
    "Joyce": "CY3190461",
    "Roger": "PD4037441",
}

CHR_LENS = [
    249250621, 243199373, 198022430, 191154276, 180915260, 171115067,
    159138663, 146364022, 141213431, 135534747, 135006516, 133851895,
    115169878, 107349540, 102531392, 90354753, 81195210, 78077248,
    59128983, 63025520, 48129895, 51304566, 155270560,
]

# Labels to export (skip unknowns)
EXPORTABLE = {"PP", "PM", "Ma", "Mb", "M1", "M2"}


def extract_column_bounds(ws, chr_len):
    for r in range(1, 200):
        if ws.cell(r, 7).value == "Column Width":
            w, col = [], 8
            while True:
                v = ws.cell(r, col).value
                if v is None:
                    break
                w.append(v)
                col += 1
            if not w:
                return None
            c = [0]
            for x in w:
                c.append(c[-1] + x)
            t = c[-1]
            return [(chr_len * c[i] / t, chr_len * c[i + 1] / t)
                    for i in range(len(w))]
    return None


def extract_grid_rows(ws):
    rows = {}
    for r in range(1, 200):
        if ws.cell(r, 7).value in SIBLINGS:
            rows[ws.cell(r, 7).value] = r
    return rows


def build_segments(values, bounds):
    """Build contiguous segments from a list of (label, col_start_bp, col_end_bp).
    Returns list of (label, start_bp, end_bp)."""
    segments = []
    for i, label in enumerate(values):
        if label not in EXPORTABLE:
            continue
        start_bp = int(bounds[i][0])
        end_bp = int(bounds[i][1])

        if segments and segments[-1][0] == label and segments[-1][2] == start_bp:
            # Extend previous segment (contiguous and same label)
            segments[-1] = (label, segments[-1][1], end_bp)
        else:
            segments.append((label, start_bp, end_bp))

    return segments


def process_workbook(input_path, output_path):
    print(f"Reading {input_path}...")
    wb = openpyxl.load_workbook(input_path)

    rows = []

    for chrom in range(1, 24):
        sheet_name = f"Chr{chrom}"
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        chr_len = CHR_LENS[chrom - 1]
        bounds = extract_column_bounds(ws, chr_len)
        if not bounds:
            continue

        grid_rows = extract_grid_rows(ws)
        num_cols = len(bounds)

        for sib in SIBLINGS:
            if sib not in grid_rows:
                continue

            top_row = grid_rows[sib]
            bot_row = top_row + 1

            # Read top (paternal) values
            top_vals = [ws.cell(top_row, 8 + i).value or "?" for i in range(num_cols)]
            bot_vals = [ws.cell(bot_row, 8 + i).value or "?" for i in range(num_cols)]

            # Build segments for paternal row
            for label, start, end in build_segments(top_vals, bounds):
                rows.append([chrom, sib, KITS[sib], label, start, end])

            # Build segments for maternal row
            for label, start, end in build_segments(bot_vals, bounds):
                rows.append([chrom, sib, KITS[sib], label, start, end])

    # Sort by Chr, Sibling, Grandparent, Start
    rows.sort(key=lambda r: (r[0], r[1], r[3], r[4]))

    # Write CSV
    print(f"Writing {output_path}...")
    with open(output_path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["Chr", "Sibling", "Kit", "Grandparent", "B37 Start", "B37 End"])
        for row in rows:
            writer.writerow(row)

    print(f"Exported {len(rows)} segments.")

    # Summary
    labels = {}
    for row in rows:
        labels[row[3]] = labels.get(row[3], 0) + 1
    print("  By grandparent:")
    for label, count in sorted(labels.items()):
        print(f"    {label}: {count} segments")


def main():
    if len(sys.argv) > 1 and not sys.argv[1].startswith("--"):
        input_path = sys.argv[1]
    else:
        input_path = DEFAULT_INPUT

    if not os.path.exists(input_path):
        print(f"Error: {input_path} not found.")
        print(f"Run fill_grandparent_grid.py first to generate the phased Excel.")
        sys.exit(1)

    os.makedirs(OUT_DIR, exist_ok=True)
    process_workbook(input_path, DEFAULT_OUTPUT)


if __name__ == "__main__":
    main()
