# -*- coding: utf-8 -*-
"""
fill_grandparent_grid.py — Implements phasing_algorithm.md

Two-phase approach with incremental propagation:
  Phase 1: FIR -> NIR -> HIR with immediate left/right propagation
  Phase 2: Label P1/P2 -> PP/PM, M1/M2 -> Ma/Mb using cousin data

Usage:
    python fill_grandparent_grid.py 9
    python fill_grandparent_grid.py 9 --log
    python fill_grandparent_grid.py all
"""
import sys
import os
import openpyxl
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter as gl
from itertools import combinations

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
VP_DIR = os.path.dirname(SCRIPT_DIR)
OUT_DIR = os.path.join(SCRIPT_DIR, "out")
XLSX_INPUT = os.path.join(VP_DIR, "Hickman.xlsx")
XLSX_OUTPUT = os.path.join(OUT_DIR, "Hickman_phased.xlsx")

SIBLINGS = ["Debbie", "Denny", "Donal", "Joyce", "Roger"]
PAIRS = list(combinations(SIBLINGS, 2))

CHR_LENS = [
    249250621, 243199373, 198022430, 191154276, 180915260, 171115067,
    159138663, 146364022, 141213431, 135534747, 135006516, 133851895,
    115169878, 107349540, 102531392, 90354753, 81195210, 78077248,
    59128983, 63025520, 48129895, 51304566, 155270560,
]

COUSINS = [
    {"name": "Mickey", "grandparent": "M?"},
    {"name": "Amy", "grandparent": "PM"},
]

# RP Overrides: manually remove known-bad RPs detected by VP auto-detection.
# Format: {chromosome: {sibling: {"remove": [col_index, ...]}}}
# Col index is 0-based (H=0, I=1, J=2, etc.)
# Based on expert review (Mick Jolley) of VP output.
RP_OVERRIDES = {
    9: {
        "Joyce": {"remove": [2]},  # Remove RP at col J (1-column noise)
    },
}

MATERNAL_SIDES = {"M?", "MP", "MM"}
PATERNAL_SIDES = {"P?", "PP", "PM"}

FILLS = {
    "PP": PatternFill("solid", fgColor="FF00FF"),
    "PM": PatternFill("solid", fgColor="98FF00"),
    "Ma": PatternFill("solid", fgColor="00FFFF"),
    "Mb": PatternFill("solid", fgColor="FFCC00"),
    "M1": PatternFill("solid", fgColor="00FFFF"),
    "M2": PatternFill("solid", fgColor="FFCC00"),
    "P?": PatternFill("solid", fgColor="C0C0C0"),
    "M?": PatternFill("solid", fgColor="C0C0C0"),
    "?":  PatternFill("solid", fgColor="C0C0C0"),
}


# ══════════════════════════════════════════════════════════════════
# TRACE LOG
# ══════════════════════════════════════════════════════════════════

class Log:
    def __init__(self, chrom, enabled=False):
        self.chrom = chrom
        self.enabled = enabled
        self.lines = []

    def __call__(self, msg, indent=0):
        if self.enabled:
            self.lines.append("  " * indent + msg)

    def section(self, title):
        if self.enabled:
            self.lines.append(f"\n## {title}\n")

    def subsection(self, title):
        if self.enabled:
            self.lines.append(f"\n### {title}\n")

    def save(self):
        if not self.enabled:
            return
        d = os.path.join(OUT_DIR, "logs")
        os.makedirs(d, exist_ok=True)
        p = os.path.join(d, f"chr{self.chrom}_decisions.md")
        with open(p, "w", encoding="utf-8") as f:
            f.write(f"# Chr{self.chrom} Phasing Decision Log\n\n")
            f.write("\n".join(self.lines) + "\n")
        print(f"  Trace: {p}")


log = Log(0)


# ══════════════════════════════════════════════════════════════════
# DATA EXTRACTION
# ══════════════════════════════════════════════════════════════════

def extract_column_bounds(ws, chr_len):
    for r in range(1, 200):
        if ws.cell(r, 7).value == "Column Width":
            w, col = [], 8
            while True:
                v = ws.cell(r, col).value
                if v is None: break
                w.append(v); col += 1
            if not w: return None
            c = [0]
            for x in w: c.append(c[-1] + x)
            t = c[-1]
            return [(chr_len*c[i]/t, chr_len*c[i+1]/t) for i in range(len(w))]
    return None


def extract_rps(ws, n):
    for r in range(1, 200):
        if ws.cell(r, 7).value == SIBLINGS[0]:
            return [ws.cell(r-1, 8+i).value for i in range(n)]
    return [None]*n


def extract_grid_rows(ws):
    rows = {}
    for r in range(1, 200):
        if ws.cell(r, 7).value in SIBLINGS:
            rows[ws.cell(r, 7).value] = r
    return rows


def extract_hir_fir(ws):
    hir = {f"{a}-{b}": [] for a, b in PAIRS}
    fir = {f"{a}-{b}": [] for a, b in PAIRS}
    row = 1
    while row <= ws.max_row:
        v = ws.cell(row, 2).value
        if v and isinstance(v, str) and "-" in v and not any(c["name"] in v for c in COUSINS):
            is_f = "FIR" in v
            pair = v.replace(" FIR Table", "")
            if pair not in hir: row += 1; continue
            row += 2; segs = []
            while row <= ws.max_row and ws.cell(row, 3).value is not None:
                s, e = ws.cell(row, 3).value, ws.cell(row, 4).value
                if isinstance(s, (int, float)) and isinstance(e, (int, float)):
                    segs.append((s, e))
                row += 1
            (fir if is_f else hir)[pair] = segs
        else: row += 1
    return hir, fir


def extract_cousins(ws):
    res = {c["name"]: {s: [] for s in SIBLINGS} for c in COUSINS}
    row = 1
    while row <= ws.max_row:
        v = ws.cell(row, 2).value
        if v and isinstance(v, str) and "FIR" not in v:
            for c in COUSINS:
                if f"-{c['name']}" in v:
                    sib = v.split("-")[0]
                    if sib not in SIBLINGS: break
                    row += 2
                    while row <= ws.max_row and ws.cell(row, 3).value is not None:
                        s, e = ws.cell(row, 3).value, ws.cell(row, 4).value
                        if isinstance(s, (int, float)) and isinstance(e, (int, float)):
                            res[c["name"]][sib].append((s, e))
                        row += 1
                    break
            else: row += 1
        else: row += 1
    return res


# ══════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════

def hit(segs, pos):
    return any(s <= pos <= e for s, e in segs)

def range_hit(segs, start, end):
    return any(s < end and e > start for s, e in segs)

def pair_key(a, b):
    for p in PAIRS:
        if set(p) == {a, b}: return f"{p[0]}-{p[1]}"
    return None

def col_mid(bounds, i):
    return (bounds[i][0] + bounds[i][1]) / 2

def pair_at(pos, pk, hir, fir):
    if hit(fir.get(pk, []), pos): return "FIR"
    if hit(hir.get(pk, []), pos): return "HIR"
    return "NIR"

def flip_p(v):
    return "P2" if v == "P1" else ("P1" if v == "P2" else "P?")

def flip_m(v):
    return "M2" if v == "M1" else ("M1" if v == "M2" else "M?")


# ══════════════════════════════════════════════════════════════════
# COUSIN RECOMBINATION FILTERING
# ══════════════════════════════════════════════════════════════════

def filter_cousin_rps(csegs, sib_rps, bounds, nc):
    log.section("Cousin Recombination Filtering")
    TOL = 2_000_000

    rp_pos = set()
    for s in SIBLINGS:
        for rc in sib_rps[s]:
            if rc < nc: rp_pos.add(bounds[rc][1])

    for c in COUSINS:
        cn = c["name"]
        all_b = set()
        for s in SIBLINGS:
            for st, en in csegs[cn][s]:
                all_b.add(st); all_b.add(en)
        if not all_b: continue

        crps = []
        for b in sorted(all_b):
            n = sum(1 for s in SIBLINGS
                    if any(abs(st-b) <= TOL or abs(en-b) <= TOL for st, en in csegs[cn][s]))
            near = n <= 2 and any(abs(b-rp) <= TOL for rp in rp_pos)
            if n >= 3 and not near:
                if not any(abs(b-cr) <= TOL for cr in crps):
                    crps.append(b)
                    log(f"{cn} boundary at {b/1e6:.1f} Mb: {n}/5 siblings -> cousin RP")

        if not crps: continue

        for s in SIBLINGS:
            segs = sorted(csegs[cn][s])
            if len(segs) < 2: continue
            srps = set(bounds[rc][1] for rc in sib_rps[s] if rc < nc)
            merged = [segs[0]]
            for seg in segs[1:]:
                gap_crp = any(merged[-1][1]-TOL <= cr <= seg[0]+TOL for cr in crps)
                gap_srp = any(merged[-1][1]-TOL <= rp <= seg[0]+TOL for rp in srps)
                if gap_crp and not gap_srp:
                    log(f"  {s}: merged ({merged[-1][0]/1e6:.1f}-{merged[-1][1]/1e6:.1f}) + ({seg[0]/1e6:.1f}-{seg[1]/1e6:.1f})")
                    merged[-1] = (merged[-1][0], seg[1])
                elif gap_crp and gap_srp:
                    log(f"  {s}: kept separate ({merged[-1][1]/1e6:.1f}-{seg[0]/1e6:.1f}) [sibling RP]")
                    merged.append(seg)
                else:
                    merged.append(seg)
            csegs[cn][s] = merged

    return csegs


# ══════════════════════════════════════════════════════════════════
# RP CLASSIFICATION
# ══════════════════════════════════════════════════════════════════

def classify_rp(sib, rp_col, bounds, csegs, nc):
    """Returns 'top', 'bottom', or 'unknown'."""
    if rp_col + 1 >= nc: return "unknown"
    bp = col_mid(bounds, rp_col)
    ap = col_mid(bounds, rp_col + 1)

    for c in COUSINS:
        segs = csegs[c["name"]][sib]
        if not segs: continue
        b, a = hit(segs, bp), hit(segs, ap)
        gp = c["grandparent"]
        if gp in MATERNAL_SIDES:
            return "bottom" if b != a else "top"
        elif gp in PATERNAL_SIDES:
            return "top" if b != a else "bottom"
    return "unknown"


# ══════════════════════════════════════════════════════════════════
# CORE: PROPAGATE LEFT/RIGHT FROM A CELL
# ══════════════════════════════════════════════════════════════════

def propagate(grid, sib, col, sib_rps, bounds, csegs, nc):
    """From grid[sib][col] (which must be set), propagate left and right
    within this sibling's segments, flipping at RPs. Returns list of
    newly set columns."""
    rp_set = set(sib_rps[sib])
    top, bot = grid[sib][col]
    newly_set = []

    # Walk RIGHT
    t, b = top, bot
    for i in range(col + 1, nc):
        if grid[sib][i] is not None:
            break  # already filled — stop
        # Check if previous col has RP (right edge)
        if i - 1 in rp_set:
            which = classify_rp(sib, i-1, bounds, csegs, nc)
            old_t, old_b = t, b
            if which == "top": t = flip_p(t)
            elif which == "bottom": b = flip_m(b)
            else: t = "P?"
            log(f"  -> {gl(8+i)}: RP at {gl(8+i-1)} right edge, {which} changed: ({old_t},{old_b})->({t},{b})", 2)
        grid[sib][i] = (t, b)
        newly_set.append(i)

    # Walk LEFT
    t, b = top, bot
    for i in range(col - 1, -1, -1):
        if grid[sib][i] is not None:
            break
        # Check if col i has RP (right edge of col i = boundary between i and i+1)
        if i in rp_set:
            which = classify_rp(sib, i, bounds, csegs, nc)
            old_t, old_b = t, b
            if which == "top": t = flip_p(t)
            elif which == "bottom": b = flip_m(b)
            else: t = "P?"
            log(f"  ← {gl(8+i)}: RP at {gl(8+i)} right edge, {which} changed: ({old_t},{old_b})->({t},{b})", 2)
        grid[sib][i] = (t, b)
        newly_set.append(i)

    return newly_set


# ══════════════════════════════════════════════════════════════════
# PHASE 1: STRUCTURAL PHASING
# ══════════════════════════════════════════════════════════════════

def phase1(anchor, sib_rps, bounds, csegs, hir, fir, nc):
    """Fill entire grid with P1/P2/M1/M2."""
    grid = {s: [None] * nc for s in SIBLINGS}

    # ── Step 2: Initialize anchor ─────────────────────────────
    log.section(f"Step 2: Initialize Anchor ({anchor})")

    # Find longest segment, pick middle column
    segments = []
    rps = sorted(sib_rps[anchor])
    breaks = sorted(set([0] + [r+1 for r in rps] + [nc]))
    for i in range(len(breaks)-1):
        if breaks[i] < breaks[i+1]:
            segments.append(list(range(breaks[i], breaks[i+1])))

    longest = max(segments, key=len)
    mid_col = longest[len(longest) // 2]
    grid[anchor][mid_col] = ("P1", "M1")
    log(f"Longest segment: cols {gl(8+longest[0])}-{gl(8+longest[-1])}")
    log(f"Set {gl(8+mid_col)} = (P1, M1)")

    newly = propagate(grid, anchor, mid_col, sib_rps, bounds, csegs, nc)
    log(f"Propagated {anchor}: {len(newly)+1} columns filled")

    # ── Step 3: FIR Phase ─────────────────────────────────────
    def run_fir():
        log.subsection("FIR Phase")
        total_new = 0
        changed = True
        while changed:
            changed = False
            for col_i in range(nc):
                for sib in SIBLINGS:
                    if grid[sib][col_i] is not None:
                        # Check all other siblings for FIR
                        for other in SIBLINGS:
                            if other == sib or grid[other][col_i] is not None:
                                continue
                            pk = pair_key(sib, other)
                            if not pk: continue
                            mid = col_mid(bounds, col_i)
                            if pair_at(mid, pk, hir, fir) == "FIR":
                                # Copy values
                                grid[other][col_i] = grid[sib][col_i]
                                log(f"FIR: {gl(8+col_i)} {sib}->{other}: {grid[sib][col_i]}")
                                newly = propagate(grid, other, col_i, sib_rps, bounds, csegs, nc)
                                log(f"  Propagated {other}: {len(newly)+1} cols")
                                total_new += len(newly) + 1
                                changed = True
        return total_new

    # ── Step 4: NIR Phase ─────────────────────────────────────
    def run_nir():
        log.subsection("NIR Phase")
        total_new = 0
        changed = True
        while changed:
            changed = False
            for col_i in range(nc):
                for sib in SIBLINGS:
                    if grid[sib][col_i] is not None:
                        continue  # already filled
                    for other in SIBLINGS:
                        if other == sib or grid[other][col_i] is None:
                            continue
                        ot, ob = grid[other][col_i]
                        if ot in ("P?", "?") or ob in ("M?", "?"):
                            continue
                        pk = pair_key(sib, other)
                        if not pk: continue
                        mid = col_mid(bounds, col_i)
                        if pair_at(mid, pk, hir, fir) == "NIR":
                            grid[sib][col_i] = (flip_p(ot), flip_m(ob))
                            log(f"NIR: {gl(8+col_i)} {other}->{sib}: ({ot},{ob})->{grid[sib][col_i]}")
                            newly = propagate(grid, sib, col_i, sib_rps, bounds, csegs, nc)
                            log(f"  Propagated {sib}: {len(newly)+1} cols")
                            total_new += len(newly) + 1
                            changed = True
                            break  # restart scan
                    if changed: break
                if changed: break
        return total_new

    # ── Step 5: HIR Phase ─────────────────────────────────────
    def run_hir():
        log.subsection("HIR Phase")
        total_new = 0
        changed = True
        while changed:
            changed = False
            for col_i in range(nc):
                for sib in SIBLINGS:
                    if grid[sib][col_i] is not None:
                        continue
                    for other in SIBLINGS:
                        if other == sib or grid[other][col_i] is None:
                            continue
                        ot, ob = grid[other][col_i]
                        if ot in ("P?", "?") or ob in ("M?", "?"):
                            continue
                        pk = pair_key(sib, other)
                        if not pk: continue
                        mid = col_mid(bounds, col_i)
                        if pair_at(mid, pk, hir, fir) != "HIR":
                            continue

                        # Use cousin to determine which side shared
                        resolved = False
                        for c in COUSINS:
                            os_ = csegs[c["name"]][other]
                            ss = csegs[c["name"]][sib]
                            om, sm = hit(os_, mid), hit(ss, mid)
                            gp = c["grandparent"]

                            if gp in MATERNAL_SIDES:
                                if om == sm:
                                    val = (flip_p(ot), ob)
                                else:
                                    val = (ot, flip_m(ob))
                                resolved = True; break
                            elif gp in PATERNAL_SIDES:
                                if om == sm:
                                    val = (ot, flip_m(ob))
                                else:
                                    val = (flip_p(ot), ob)
                                resolved = True; break

                        if not resolved:
                            continue

                        grid[sib][col_i] = val
                        log(f"HIR: {gl(8+col_i)} {other}->{sib} (cousin={c['name']}): {grid[other][col_i]}->{val}")
                        newly = propagate(grid, sib, col_i, sib_rps, bounds, csegs, nc)
                        log(f"  Propagated {sib}: {len(newly)+1} cols")
                        total_new += len(newly) + 1
                        changed = True
                        break
                    if changed: break
                if changed: break
        return total_new

    # Run phases with cascading
    run_fir()
    run_nir()
    run_fir()  # re-run after NIR
    run_hir()
    run_fir()  # re-run after HIR
    run_nir()  # re-run after HIR

    # Mark unknowns
    for s in SIBLINGS:
        for i in range(nc):
            if grid[s][i] is None:
                grid[s][i] = ("?", "?")

    return grid


# ══════════════════════════════════════════════════════════════════
# PHASE 2: LABELING
# ══════════════════════════════════════════════════════════════════

def phase2(grid, sib_rps, bounds, csegs, fir, nc):
    log.section("Phase 2: Labeling")
    labeled = {}

    # Build segments per sibling
    def get_segs(sib):
        rps = sorted(sib_rps[sib])
        breaks = sorted(set([0] + [r+1 for r in rps] + [nc]))
        return [list(range(breaks[i], breaks[i+1]))
                for i in range(len(breaks)-1) if breaks[i] < breaks[i+1]]

    for sib in SIBLINGS:
        g = grid[sib]
        segments = get_segs(sib)

        # Find paternal mapping
        pat_map = None
        for c in COUSINS:
            gp = c["grandparent"]
            if gp not in ("PP", "PM"): continue
            cs = csegs[c["name"]][sib]
            if not cs: continue
            for seg in segments:
                if any(range_hit(cs, bounds[col][0], bounds[col][1]) for col in seg):
                    lbl = g[seg[0]][0]
                    if lbl in ("P1", "P2"):
                        pat_map = (lbl, gp)
                        log(f"{sib}: {c['name']}({gp}) matches segment with {lbl} -> {lbl}={gp}")
                        break
            if pat_map: break

        # Find maternal mapping
        mat_map = None
        for c in COUSINS:
            gp = c["grandparent"]
            if gp not in ("MP", "MM"): continue
            cs = csegs[c["name"]][sib]
            if not cs: continue
            for seg in segments:
                if any(range_hit(cs, bounds[col][0], bounds[col][1]) for col in seg):
                    lbl = g[seg[0]][1]
                    if lbl in ("M1", "M2"):
                        mat_map = (lbl, "Ma" if gp == "MP" else "Mb")
                        log(f"{sib}: {c['name']}({gp}) matches segment with {lbl} -> {lbl}={mat_map[1]}")
                        break
            if mat_map: break

        if not pat_map:
            log(f"{sib}: no paternal cousin match — checking FIR partners")

        labeled[sib] = (pat_map, mat_map)

    # Cross-sibling: propagate mapping via FIR
    for sib in SIBLINGS:
        pat_map, mat_map = labeled[sib]
        if pat_map: continue
        for ref in SIBLINGS:
            if ref == sib: continue
            rp, rm = labeled[ref]
            if not rp: continue
            pk = pair_key(sib, ref)
            if not pk: continue
            # Find a FIR column
            for ss, se in fir.get(pk, []):
                mid = (ss + se) / 2
                ci = next((i for i in range(nc) if bounds[i][0] <= mid <= bounds[i][1]), None)
                if ci is None: continue
                ref_lbl = grid[ref][ci][0]
                sib_lbl = grid[sib][ci][0]
                if sib_lbl in ("P1", "P2") and ref_lbl in ("P1", "P2"):
                    # FIR: same labels. So if ref P1=PP, and sib has P1 here too -> sib P1=PP
                    ref_mapped = rp[1] if ref_lbl == rp[0] else ("PP" if rp[1] == "PM" else "PM")
                    if sib_lbl == ref_lbl:
                        pat_map = (sib_lbl, ref_mapped)
                    else:
                        other_gp = "PP" if ref_mapped == "PM" else "PM"
                        pat_map = (sib_lbl, other_gp)
                    log(f"{sib}: FIR with {ref} at {gl(8+ci)} -> {pat_map[0]}={pat_map[1]}")
                    labeled[sib] = (pat_map, mat_map)
                    break
            if labeled[sib][0]: break

    # Apply mappings
    result = {}
    for sib in SIBLINGS:
        pat_map, mat_map = labeled[sib]
        g = grid[sib]
        new_g = []
        for t, b in g:
            # Paternal
            if pat_map and t in ("P1", "P2"):
                if t == pat_map[0]:
                    nt = pat_map[1]
                else:
                    nt = "PP" if pat_map[1] == "PM" else "PM"
            elif t in ("P1", "P2"):
                nt = t  # keep structural label
            else:
                nt = t

            # Maternal
            if mat_map and b in ("M1", "M2"):
                if b == mat_map[0]:
                    nb = mat_map[1]
                else:
                    nb = "Ma" if mat_map[1] == "Mb" else "Mb"
            elif b in ("M1", "M2"):
                nb = b  # keep structural label
            else:
                nb = b

            new_g.append((nt, nb))
        result[sib] = new_g

    return result


# ══════════════════════════════════════════════════════════════════
# ORCHESTRATION
# ══════════════════════════════════════════════════════════════════

def count_filled(grids, nc):
    f = 0
    for s in SIBLINGS:
        for i in range(nc):
            t, b = grids[s][i]
            if t in ("PP", "PM", "P1", "P2"): f += 1
            if b in ("Ma", "Mb", "M1", "M2"): f += 1
    return f, len(SIBLINGS) * nc * 2


def run_with_anchor(anchor, sib_rps, bounds, csegs, hir, fir, nc):
    log.section(f"Anchor: {anchor}")
    structural = phase1(anchor, sib_rps, bounds, csegs, hir, fir, nc)
    labeled = phase2(structural, sib_rps, bounds, csegs, fir, nc)
    filled, total = count_filled(labeled, nc)
    return labeled, structural, filled, total


def process_chromosome(ws, chrom, log_enabled=False):
    global log
    log = Log(chrom, log_enabled)

    chr_len = CHR_LENS[chrom - 1]
    bounds = extract_column_bounds(ws, chr_len)
    if not bounds:
        print(f"  Chr{chrom}: No data, skipping"); return
    nc = len(bounds)
    rp_asgn = extract_rps(ws, nc)
    grid_rows = extract_grid_rows(ws)
    hir, fir = extract_hir_fir(ws)
    csegs = extract_cousins(ws)
    if not grid_rows:
        print(f"  Chr{chrom}: No grid, skipping"); return

    sib_rps = {s: [] for s in SIBLINGS}
    for i, rp in enumerate(rp_asgn):
        if rp in SIBLINGS: sib_rps[rp].append(i)

    # Apply RP overrides (manual expert corrections)
    overrides = RP_OVERRIDES.get(chrom, {})
    if overrides:
        log.section("RP Overrides")
    for s, actions in overrides.items():
        for rc in actions.get("remove", []):
            if rc in sib_rps.get(s, []):
                sib_rps[s].remove(rc)
                log(f"Override: removed {s} RP at col {gl(8+rc)} (index {rc})")
    if overrides:
        for s in overrides:
            log(f"  {s} RPs now: {[gl(8+c) for c in sib_rps[s]]}")

    # Log summary
    log.section("Data Summary")
    for i in range(nc):
        log(f"{gl(8+i)}: {bounds[i][0]/1e6:.1f}-{bounds[i][1]/1e6:.1f} Mb  RP={rp_asgn[i] or '-'}")
    log("")
    for s in SIBLINGS:
        log(f"{s}: {len(sib_rps[s])} RPs at {[gl(8+c) for c in sib_rps[s]]}")

    csegs = filter_cousin_rps(csegs, sib_rps, bounds, nc)

    sibs_sorted = sorted(SIBLINGS, key=lambda s: len(sib_rps[s]))
    best_lab = best_struct = None
    best_filled = -1
    best_anchor = None

    for attempt in range(min(3, len(sibs_sorted))):
        anchor = sibs_sorted[attempt]
        lab, struct, filled, total = run_with_anchor(anchor, sib_rps, bounds, csegs, hir, fir, nc)
        print(f"  Chr{chrom}: Anchor={anchor} ({len(sib_rps[anchor])} RPs) -> {filled}/{total}")
        if filled > best_filled:
            best_filled, best_lab, best_struct, best_anchor = filled, lab, struct, anchor
        if filled == total: break

    print(f"  Chr{chrom}: Best = {best_anchor} ({best_filled}/{len(SIBLINGS)*nc*2})")

    # Write Excel
    for s in SIBLINGS:
        if s not in grid_rows: continue
        tr, br = grid_rows[s], grid_rows[s] + 1
        for i in range(nc):
            tv, bv = best_lab[s][i]
            ws.cell(tr, 8+i).value = tv
            ws.cell(tr, 8+i).alignment = Alignment(horizontal="center")
            ws.cell(tr, 8+i).fill = FILLS.get(tv, FILLS["?"])
            ws.cell(br, 8+i).value = bv
            ws.cell(br, 8+i).alignment = Alignment(horizontal="center")
            ws.cell(br, 8+i).fill = FILLS.get(bv, FILLS["?"])

    # Print grid
    print(f"\n  {'':>12}", end="")
    for i in range(nc): print(f"{gl(8+i):>4}", end="")
    print()
    print(f"  {'RP':>12}", end="")
    for i in range(nc): print(f"{(rp_asgn[i] or '')[:3]:>4}", end="")
    print()
    print("  " + "-" * (12 + nc * 4))
    for s in SIBLINGS:
        g = best_lab[s]
        print(f"  {s+' top':>12}", end="")
        for i in range(nc): print(f"{g[i][0]:>4}", end="")
        print(f"  RPs:{[gl(8+c) for c in sib_rps[s]]}")
        print(f"  {s+' bot':>12}", end="")
        for i in range(nc): print(f"{g[i][1]:>4}", end="")
        print(); print()

    log.save()


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    log_on = "--log" in sys.argv
    if not args:
        print("Usage: python fill_grandparent_grid.py <chr|all> [--log]"); sys.exit(1)
    chroms = list(range(1, 24)) if args[0].lower() == "all" else [int(x) for x in args]
    os.makedirs(OUT_DIR, exist_ok=True)
    print(f"Loading {XLSX_INPUT}...")
    wb = openpyxl.load_workbook(XLSX_INPUT)
    for ch in chroms:
        sheet = f"Chr{ch}"
        if sheet not in wb.sheetnames: print(f"  Chr{ch}: Not found"); continue
        process_chromosome(wb[sheet], ch, log_on)
    print(f"\nSaving {XLSX_OUTPUT}...")
    wb.save(XLSX_OUTPUT)
    print("Done.")


if __name__ == "__main__":
    main()
