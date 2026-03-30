# -*- coding: utf-8 -*-
"""
Visual_Phaser_New_V20.py performs comparisons between siblings and cousins and 
stores the results in a .xlsx file.

© 2026 Mick Jolley (mickj1948@gmail.com) 

"""
import numpy as np
import pandas as pd
import sys
from itertools import combinations
import os
from PIL import Image, ImageDraw, ImageFont
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, Alignment, PatternFill, Font
from openpyxl.utils.cell import column_index_from_string as cs
from openpyxl.utils import get_column_letter as cl
import time
import platform
from collections import Counter

from VPnew_config_V20 import *

def conditions(al1x, al2x, al1y, al2y):
    if al1x == NO_CALL or al1y == NO_CALL:
        return 'limegreen'
    elif al1x == al2x and al1y == al2y and al1x != al1y:
        return 'crimson'
    elif (al1x == al1y and al2x == al2y) or (al1x == al2y and al2x == al1y):
        return 'limegreen'
    else:
        return 'yellow'

def load_dna_files(pair, chrom):
    file_names = os.listdir(FILES_PATH)

    dm = pd.DataFrame()

    for ind in list(pair):
        for filname in file_names:
            if ind + "_raw" in filname:
                file_name = filname

                this_file = os.path.join(FILES_PATH, file_name)

                # Read the different types of files
                if "Ancestry" in file_name:
                    df = pd.read_csv(
                        this_file,
                        sep="\t",
                        skip_blank_lines=True,
                        comment="#",
                        header=0,
                        names=[
                            "rsid",
                            "chromosome",
                            "position",
                            f"{ind}-allele1",
                            f"{ind}-allele2",
                        ],
                    )
                else:
                    df = pd.read_csv(
                        this_file,
                        sep="\t",
                        skip_blank_lines=True,
                        comment="#",
                        header=0,
                        low_memory=False,
                        names=[
                            "rsid",
                            "chromosome",
                            "position",
                            f"{ind}_allele_pair",
                        ],
                    )

                    df[f"{ind}-allele1"] = df[f"{ind}_allele_pair"].str[0]
                    df[f"{ind}-allele2"] = df[f"{ind}_allele_pair"].str[1]

                    # df.drop(columns=['allele_pair'])
                    df.drop([f"{ind}_allele_pair"], axis=1, inplace=True)

                    # Drop Y and MT chromosomes
                    df.replace("X", "23", inplace=True)
                    df.replace("XY", "23", inplace=True)
                    df = df.drop(df[df["chromosome"] == "Y"].index)
                    df = df.drop(df[df["chromosome"] == "MT"].index)

                    # Set "chromosome" dtype to 'int64'
                    df = df.astype({"chromosome": "int64"})

                df = df[df["chromosome"] == chrom]
                
                df = df[df[f"{ind}-allele1"].eq("A") 
                    | df[f"{ind}-allele1"].eq("T")
                    | df[f"{ind}-allele1"].eq("C")
                    | df[f"{ind}-allele1"].eq("G")
                    | df[f"{ind}-allele1"].eq(NO_CALL)
                ]
                
                df = df.sort_values(by='position')

                df = df.reset_index(drop=True)

                # Merge this person's DNA dataframe into a common dataframe.
                if len(dm) > 0:
                    dm = pd.merge(dm, df, on=("rsid", "chromosome", "position"))
                else:
                    dm = df

    return dm

def scan_genomes(dm, chrom):
    dm["match"] = np.vectorize(conditions)(
        dm[f"{pair[0]}-allele1"],
        dm[f"{pair[0]}-allele2"],
        dm[f"{pair[1]}-allele1"],
        dm[f"{pair[1]}-allele2"],
    )
    
    if REPAIR_FILES:
        dm = repair_files(dm)

    segflag = False
    stpos = 0
    pos = 0
    nmms = 0
    fflag = False
    fsnps = 0
    
    dx = pd.DataFrame()
    dplot = pd.DataFrame()
    ds = pd.DataFrame()
    
    if chrom == 23:
        cutoff = X_HIR_CUTOFF
        fcutoff = X_FIR_CUTOFF
    else:
        cutoff = HIR_CUTOFF
        fcutoff = FIR_CUTOFF
        
    length = len(dm)    
    for i in range(length):
        if i == 0 and (dm.loc[i, "match"] == 'yellow' or dm.loc[i, "match"] == 'limegreen'):
            nsnps = 1
            segflag = True
            stpos = dm.loc[i, "position"]
            if dm.loc[i,'match'] == 'limegreen':
                fsnps = 1
                fstpos = dm.loc[i, "position"]
                fflag = True
        elif not segflag and (dm.loc[i, "match"] == 'yellow' or dm.loc[i, "match"] == 'limegreen'):
            nsnps = 1
            segflag = True
            stpos = dm.loc[i, "position"]
            if not fflag and dm.loc[i,'match'] == 'limegreen':
                fsnps = 1
                fstpos = dm.loc[i, "position"]
                fflag = True
        elif not fflag and dm.loc[i,'match'] == 'limegreen':
            fsnps = 1
            fstpos = dm.loc[i, "position"]
            fflag = True
        elif segflag and (dm.loc[i, "match"] == 'yellow' or dm.loc[i, "match"] == 'limegreen'):
            nsnps += 1
            pos = dm.loc[i, "position"]
            if fflag:
                if dm.loc[i,'match'] == 'limegreen':
                    fsnps += 1
                    fpos = dm.loc[i, "position"]
                else: 
                    fflag = False
                    if fsnps > FIR_SNP_MIN:
                        # Calculate length (cM).
                        dcm = cm_calc(fstpos, fpos)
    
                        if dcm > fcutoff:
                            addn = pd.DataFrame(
                                {
                                    "Chr": chrom,
                                    "Start Mb": fstpos,
                                    "Finish Mb": fpos,
                                    "No. SNPs": fsnps,
                                    "Length (cM)": round(dcm, 1),
                                },
                                index=[0],
                            )
                            ds = pd.concat([ds, addn], ignore_index=True)
                    fsnps = 0
        elif segflag and dm.loc[i, "match"] == 'crimson':
            fflag = False
            if fsnps > FIR_SNP_MIN:
                # Calculate length (cM).
                dcm = cm_calc(fstpos, fpos)

                if dcm > fcutoff:
                    addn = pd.DataFrame(
                        {
                            "Chr": chrom,
                            "Start Mb": fstpos,
                            "Finish Mb": fpos,
                            "No. SNPs": fsnps,
                            "Length (cM)": round(dcm, 1),
                        },
                        index=[0],
                    )
                    ds = pd.concat([ds, addn], ignore_index=True)
                    
            fsnps = 0
            
            nmms += 1
            if nmms == 1:
                mmpos = dm.loc[i, "position"]
            else:
                if dm.loc[i, "position"] - mmpos < MM_DIST * 1000:
                    segflag = False
                    nmms = 0

                    if nsnps > HIR_SNP_MIN:
                        # Calculate length (cM).
                        dcm = cm_calc(stpos, pos)

                        if dcm > cutoff:
                            addn = pd.DataFrame(
                                {
                                    "Chr": chrom,
                                    "Start Mb": stpos,
                                    "Finish Mb": pos,
                                    "No. SNPs": nsnps,
                                    "Length (cM)": round(dcm, 1),
                                },
                                index=[0],
                            )
                            dx = pd.concat([dx, addn], ignore_index=True)
                            
                    nsnps = 0
                                        
                else:
                    nmms = 1
                    mmpos = dm.loc[i, "position"]

    if i == len(dm) - 1:
        if nsnps > HIR_SNP_MIN:
            # Calculate length (cM).
            dcm = cm_calc(stpos, pos)

            if dcm > cutoff:
                addn = pd.DataFrame(
                    {
                        "Chr": chrom,
                        "Start Mb": stpos,
                        "Finish Mb": pos,
                        "No. SNPs": nsnps,
                        "Length (cM)": round(dcm, 1),
                    },
                    index=[0],
                )
                dx = pd.concat([dx, addn], ignore_index=True)
                
            if fsnps > FIR_SNP_MIN:
                # Calculate length (cM).
                dcm = cm_calc(fstpos, fpos)

                if dcm > fcutoff:
                    addn = pd.DataFrame(
                        {
                            "Chr": chrom,
                            "Start Mb": fstpos,
                            "Finish Mb": fpos,
                            "No. SNPs": fsnps,
                            "Length (cM)": round(dcm, 1),
                        },
                        index=[0],
                    )
                    ds = pd.concat([ds, addn], ignore_index=True)
    
    return dx, ds

def get_image(dplot, pair_name, chrom):

    img = Image.new("RGB", (len(dplot), 35), color="white")

    # create line image
    img1 = ImageDraw.Draw(img)

    pos = 0

    for i in range(len(dplot)):
        color = dplot.loc[i,'match']
        color2 = dplot.loc[i,'bar']
        img1.line([(pos, 0), (pos, 19)], fill=color, width=0)
        img1.line([(pos, 20), (pos, 34)], fill=color2, width=0)
        pos += 1
        
    img.save(wdir + f'{pair_name} {chrom}.png')  
    
def get_scale(dplot, chrom):
    
    img = Image.new("RGB", (len(dplot) + 30, 35), color="white")
    
    img1 = ImageDraw.Draw(img)
    
    if platform.system() == 'Windows':    
        fnt = ImageFont.truetype("arial.ttf", 13)
        fnt1 = ImageFont.truetype("arial.ttf", 10)
    elif platform.system() == 'Darwin':    
        fnt = ImageFont.truetype("/System/Library/Fonts/Supplemental/Arial.ttf", 13)    
        fnt1 = ImageFont.truetype("/System/Library/Fonts/Supplemental/Arial.ttf", 10)
    elif platform.system() == 'Linux':
        fnt = ImageFont.truetype(LINUX_FONT_STRING, 13)
        fnt1 = ImageFont.truetype(LINUX_FONT_STRING, 10)
        
    base_position = 0
    
    for SNP in list(dplot['position']):
            
                if base_position % 50 == 0:
                    img1.text(( base_position,5),f"{SNP / 1000000:0.1f}\n|", 
                              font=fnt, fill="black")
                    
                elif base_position % 5 == 0:
                    img1.text((base_position, 21), '|', font=fnt1, fill="black")
                    
                base_position +=1
                
    img.save(wdir + f'scale {chrom}.png')
    
def load_map():
    file_name = os.path.join(MAP_PATH, "min_map.txt")

    dmap_source = pd.read_csv(file_name, sep="\t", header=0)

    return dmap_source

def cm_calc(st, fin):
    stcm = 0
    fincm = dmap.loc[len(dmap) - 1, "cM"]

    for w in range(1, len(dmap)):
        if st >= dmap.loc[w - 1, "Position"] and st < dmap.loc[w, "Position"]:
            stcm = dmap.loc[w - 1, "cM"] + (
                (st - dmap.loc[w - 1, "Position"])
                / (dmap.loc[w, "Position"] - dmap.loc[w - 1, "Position"])
            ) * (dmap.loc[w, "cM"] - dmap.loc[w - 1, "cM"])

        if fin > dmap.loc[w - 1, "Position"] and fin <= dmap.loc[w, "Position"]:
            fincm = dmap.loc[w - 1, "cM"] + (
                (fin - dmap.loc[w - 1, "Position"])
                / (dmap.loc[w, "Position"] - dmap.loc[w - 1, "Position"])
            ) * (dmap.loc[w, "cM"] - dmap.loc[w - 1, "cM"])

            break

    dcm = fincm - stcm

    return dcm

def paste_image(pair_name, chrom, q, dxtot):
            
    if q == 0 and COUSINS == []:

        if SCALE_ON:
            img = openpyxl.drawing.image.Image(wdir + f'scale {chrom}.png')
            ws.add_image(img, ws.cell(1, 8).coordinate)
            
    if not SHOW_NO_MATCHES:
        
        if COUSINS  !=  []:
            
            # Set names column width
            if len(pair_name) > ws.column_dimensions["G"].width:
                ws.column_dimensions["G"].width = len(pair_name) + 2
        
        if pair_name in list(dxtot['pair']):
            
            # Set names column width
            if len(pair_name) > ws.column_dimensions["G"].width:
                ws.column_dimensions["G"].width = len(pair_name) + 2
                 
            img = openpyxl.drawing.image.Image(wdir + f'{pair_name} {chrom}.png')
                            
            if COUSINS  !=  []:
                next_line = find_next_line(ws,7,3)
            else:            
                next_line = find_next_line(ws,7,2)            
                    
            ws.add_image(img, ws.cell(next_line, 8).coordinate)
            ws.cell(next_line, 7).value = pair_name
            ws.cell(next_line, 7).alignment = Alignment(horizontal="center")
    else:
        
        # Set names column width
        if len(pair_name) > ws.column_dimensions["G"].width:
            ws.column_dimensions["G"].width = len(pair_name) + 2
                        
        img = openpyxl.drawing.image.Image(wdir + f'{pair_name} {chrom}.png')
        
        if COUSINS  !=  []: 
            img.width = img.width * im_width/len(dplot)
                
        if q == 0:
            if COUSINS  !=  []:
                next_line = find_next_line(ws,7,3) 
            else:    
                next_line = 3
        else:            
            next_line = find_next_line(ws,7,2)            
                
        ws.add_image(img, ws.cell(next_line, 8).coordinate)
        ws.cell(next_line, 7).value = pair_name
        ws.cell(next_line, 7).alignment = Alignment(horizontal="center")
                                
def delete_images():
    files = os.listdir(wdir)
    for images in files:
        if images.endswith(".png"):
            os.remove(wdir + images)
            
def format_sheet(ws):
    
    ws.column_dimensions["A"].width = 1
    ws.column_dimensions["B"].width = 5
    ws.column_dimensions["C"].width = 11
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 11
    ws.column_dimensions["F"].width = 13
    ws.column_dimensions["G"].width = 14
    
def add_borders(ws, col):
    
    next_line = find_next_line(ws,7,3)
    
    side = Side(border_style="thick")
    border = Border(left=side)
    for i in range(1, next_line):
        for j in range(8, col+1):
            ws.cell(i, j).border = border
            
def add_chroms(ws, col):
    if AUTO_REC_PNTS:
        next_line = find_next_line(ws,7,-len(SIBLINGS)*3+1) 
    else:
        next_line = find_next_line(ws,7,3)
        
    for w, name in enumerate(SIBLINGS):
        ws.cell(next_line + w*3, 7).value = name
        ws.cell(next_line + w*3, 7).alignment = Alignment(horizontal="center")

        for i in range(8, col):
            ws.cell(next_line + w*3, i).fill = PatternFill(
                "solid", fgColor="FF00FF"
            )
            ws.cell(next_line + 1 + w*3, i).fill = PatternFill(
                "solid", fgColor="98FF00"
            )
            
    ws.cell(3, col+1).fill = PatternFill(
        "solid", fgColor="FF00FF"
    )
    ws.cell(4, col+1).fill = PatternFill(
        "solid", fgColor="98FF00"
    )
    ws.cell(6, col+1).fill = PatternFill(
        "solid", fgColor="00FFFF"
    )
    ws.cell(7, col+1).fill = PatternFill(
        "solid", fgColor="FFCC00"
    )
    ws.cell(9, col+1).fill = PatternFill(
        "solid", fgColor="FF00FF"
    )
    ws.cell(10, col+1).fill = PatternFill(
        "solid", fgColor="FFCC00"
    )
    ws.cell(12, col+1).fill = PatternFill(
        "solid", fgColor="00FFFF"
    )
    ws.cell(13, col+1).fill = PatternFill(
        "solid", fgColor="98FF00"
    )
                        

def find_next_line(ws, col, addn):
    
    lr = 0
    for i in range(1,1000):
        if ws.cell(i,col).value != None:
            lr = i
    next_line = lr + addn 
    
    return next_line

def get_dplot(q, dtot, dxtot, dstot, pair_name):
    global CHROM_TRUE_SIZE, RESOLUTION

    if LINEAR_CHROMOSOME:
        CHROM_TRUE_SIZE = False
        if RESOLUTION != 10:
            RESOLUTION = 1    
                    
    res = RESOLUTION * 1000
    if res > len(dtot):
        res = len(dtot)
        
    if CHROM_TRUE_SIZE:
        length = dtot.loc[len(dtot)-1,'position'] - dtot.loc[0,'position']
        div = (len(dtot) * 250000000)//(res * length)
    else:
        div =  len(dtot)//res 
            
    dplot = pd.DataFrame(data={'match':'grey'},index=np.arange(len(dtot)//div + 1)) 
    
    ycnt = 0
    rcnt = 0
    gcnt = 0
    gycnt = 0
         
    for i in range(len(dtot)):
        if dtot.iloc[i,q+1] == 'crimson':
            rcnt += 1
        elif dtot.iloc[i,q+1] == 'yellow':
            ycnt += 1
        elif dtot.iloc[i,q+1] == 'limegreen':
            gcnt += 1
        elif dtot.iloc[i,q+1] == 'grey':
            gycnt += 1

        if i % div == 0 and i != 0:
            if rcnt > 0:
                dplot.loc[int(i/div) - 1,'match'] = 'crimson'
                dplot.loc[int(i/div) - 1,'position'] = dtot.loc[i,'position']
            elif ycnt > gycnt:
                dplot.loc[int(i/div) - 1,'match'] = 'yellow'
                dplot.loc[int(i/div) - 1,'position'] = dtot.loc[i,'position']
            elif gcnt > gycnt:
                dplot.loc[int(i/div) - 1,'match'] = 'limegreen'
                dplot.loc[int(i/div) - 1,'position'] = dtot.loc[i,'position']
            else:
                dplot.loc[int(i/div) - 1,'position'] = dtot.loc[i,'position']

            ycnt = 0
            rcnt = 0
            gcnt = 0
            gycnt = 0
            
    if rcnt > 0:
        dplot.loc[len(dplot)-1,'match'] = 'crimson'
        dplot.loc[len(dplot)-1,'position'] = dtot.loc[len(dtot)-1,'position']

    elif ycnt > gycnt:
        dplot.loc[len(dplot)-1,'match'] = 'yellow'
        dplot.loc[len(dplot)-1,'position'] = dtot.loc[len(dtot)-1,'position']

    elif gcnt > gycnt:
        dplot.loc[len(dplot)-1,'match'] = 'limegreen'
        dplot.loc[len(dplot)-1,'position'] = dtot.loc[len(dtot)-1,'position']

    else:
        dplot.loc[len(dplot)-1,'position'] = dtot.loc[len(dtot)-1,'position']
    
    dplot['bar'] = 'black'
    
    
    for i in range(len(dplot)):
        for j in range(len(dxtot)):
            if dxtot.loc[j,'pair'] == pair_name:
                if dxtot.loc[j,'Start Mb'] <= dplot.loc[i,'position'] <= dxtot.loc[j,'Finish Mb']:
                    dplot.loc[i,'bar'] = 'blue'

        for j in range(len(dstot)):
            if dstot.loc[j,'pair'] == pair_name:
                if dstot.loc[j,'Start Mb'] <= dplot.loc[i,'position'] <= dstot.loc[j,'Finish Mb']:
                    dplot.loc[i,'bar'] = 'orange'
        if i != 0:                                           
            if dplot.loc[i,'bar'] != dplot.loc[i-1,'bar']:
                if pair_name.split('-')[0] in SIBLINGS and pair_name.split('-')[1] in SIBLINGS:
                    rps_list.append(i)
                    rnames_list.append(pair_name)
                                            
    if LINEAR_CHROMOSOME:
        
        length = chr_lens[chrom - 1] 
        
        dplot = dplot.dropna(ignore_index=True)
        
        if RESOLUTION == 10:
            dtot['fract'] = 10000 * round(dtot['position']/length,4)
            dplotr = pd.DataFrame(data={'match':'grey','bar':'grey'},index=np.arange(10001)) 
            dplotr['position'] = np.linspace(0,length,10001)    
        else:
            dtot['fract'] = 1000 * round(dtot['position']/length,3)
            dplotr = pd.DataFrame(data={'match':'grey','bar':'grey'},index=np.arange(1001)) 
            dplotr['position'] = np.linspace(0,length,1001)    
            
        dtot = dtot.astype({"fract": "int64"})
        fdict = dict(zip(dtot['position'],dtot['fract']))
                
        for i in range(len(dplot)):
            f = fdict[dplot.loc[i,'position']]
            dplotr.loc[f,'match'] = dplot.loc[i,'match']
            dplotr.loc[f,'bar'] = dplot.loc[i,'bar']
                        
        dplot = dplotr.copy()
        
    return dplot, rps_list, rnames_list

def paste_tables(dx, ds, pair_name):
    # Cell formatting
    side = Side(border_style="thin")
    border = Border(left=side, right=side, top=side, bottom=side)
    align = Alignment(horizontal="center")
        
    if not SHOW_NO_MATCHES:
     
        if len(dx) > 0:
                        
            next_line = find_next_line(ws,2,2)        
            
            ws.cell(next_line, 2).value = pair_name
            
            title = list(dx.columns)
            
            for i,item in enumerate(title):
                ws.cell(next_line + 1, 2 + i ).value = item
                ws.cell(next_line + 1, 2 + i).alignment = align
                ws.cell(next_line + 1, 2 + i).border = border
                ws.cell(next_line + 1, 2 + i).alignment = align
            for i in range(len(dx)):
                for j in range(5):
                    ws.cell(next_line + 2 + i, 2 + j ).value = dx.iloc[i,j]
                    ws.cell(next_line + 2 + i, 2 + j).alignment = align
                    ws.cell(next_line + 2 + i, 2 + j).border = border

            if FIR_TABLES:
                
                if len(ds) > 0:
                    
                    next_line = find_next_line(ws,2,2)  
                    
                    ws.cell(next_line, 2).value = pair_name + ' FIR Table'
                    
                    title = list(ds.columns)
                    
                    for i,item in enumerate(title):
                        ws.cell(next_line + 1, 2 + i ).value = item
                        ws.cell(next_line + 1, 2 + i).alignment = align
                        ws.cell(next_line + 1, 2 + i).border = border
                        ws.cell(next_line + 1, 2 + i).alignment = align
                    for i in range(len(ds)):
                        for j in range(5):
                            ws.cell(next_line + 2 + i, 2 + j ).value = ds.iloc[i,j]
                            ws.cell(next_line + 2 + i, 2 + j).alignment = align
                            ws.cell(next_line + 2 + i, 2 + j).border = border
            
    else:
        
        next_line = find_next_line(ws,2,2)        
        
        if len(dx) > 0:
        
            ws.cell(next_line, 2).value = pair_name
            
            title = list(dx.columns)
            
            for i,item in enumerate(title):
                ws.cell(next_line + 1, 2 + i ).value = item
                ws.cell(next_line + 1, 2 + i).alignment = align
                ws.cell(next_line + 1, 2 + i).border = border
                ws.cell(next_line + 1, 2 + i).alignment = align
            for i in range(len(dx)):
                for j in range(5):
                    ws.cell(next_line + 2 + i, 2 + j ).value = dx.iloc[i,j]
                    ws.cell(next_line + 2 + i, 2 + j).alignment = align
                    ws.cell(next_line + 2 + i, 2 + j).border = border 
                    
            if FIR_TABLES:        
                    
                if len(ds) > 0:
                    
                    next_line = find_next_line(ws,2,2)  
                    
                    ws.cell(next_line, 2).value = pair_name + ' FIR Table'
                    
                    title = list(ds.columns)
                    
                    for i,item in enumerate(title):
                        ws.cell(next_line + 1, 2 + i ).value = item
                        ws.cell(next_line + 1, 2 + i).alignment = align
                        ws.cell(next_line + 1, 2 + i).border = border
                        ws.cell(next_line + 1, 2 + i).alignment = align
                    for i in range(len(ds)):
                        for j in range(5):
                            ws.cell(next_line + 2 + i, 2 + j ).value = ds.iloc[i,j]
                            ws.cell(next_line + 2 + i, 2 + j).alignment = align
                            ws.cell(next_line + 2 + i, 2 + j).border = border            

def do_arp(rps_list, rnames_list):
    
    rcomb_list = list(zip(rps_list, rnames_list))
    
    side = Side(border_style="thick")
    border = Border(left=side)
    
    next_line = find_next_line(ws,7,3) 
    arp_row = next_line + len(SIBLINGS) * 3

    # Remove duplicates and sort.
    rps_list = list(set(rps_list)) 
    rps_list.sort()
    
    # Add beginning and end pixels.
    rps_list.append(len(dplot)-1)
    rps_list.insert(0,0)
    
    rpsf = []
    
    arp_tol = ARP_TOLERANCE * RESOLUTION
                
    # Remove RPs within a distance of tol of each other.    
    for num in rps_list:
         if not any(abs(num - x) < arp_tol for x in rpsf):
             rpsf.append(num)

    # Make columns.
    rpsf1 = rpsf[1:]
    rpsf2 = rpsf[:-1]
    rpsf_diff = np.subtract(rpsf1, rpsf2)
    
    # Set column widths.
    for i, rp in enumerate(rpsf_diff):
        ws.column_dimensions[cl(8 + i)].width = rp * SCALE_FACTOR
        ws.cell(arp_row, 8 + i).value = rp
        ws.cell(arp_row, 8 + i).alignment = Alignment(horizontal='center')

        ws.cell(arp_row, 7).value = 'Column Width'
        ws.cell(arp_row, 7).alignment = Alignment(horizontal='center')

    if AUTO_RP_ASSIGN and len(SIBLINGS) != 2:
        
        # Make match pairs for each recombination point.
        rpasstot = []
            
        for rp in rpsf[1:-1]:
            rpass = []
            for rn in rcomb_list:
                if abs(rn[0] - rp) < arp_tol:
                    rpass.append((rn[1].split('-')[0], rn[1].split('-')[1]))
            rpasstot.append(rpass)    
        
        # Find assigments.
        for i,item in enumerate(rpasstot):
            temp = []
            for el in item:
                for s in el:
                    temp.append(s)
            res = Counter(temp)        
            
            keymax = max(zip(res.values(), res.keys()))[1] 
                        
            ws.cell(arp_row - 3 * len(SIBLINGS), 8 + i).value = keymax
            ws.cell(arp_row - 3 * len(SIBLINGS), 8 + i).alignment = Alignment(horizontal='right')
            ws.cell(arp_row - 3 * len(SIBLINGS), 8 + i).font = Font(bold=True)

    return len(rpsf_diff) + 8

def repair_files(dm):
    firs = FIR_SNP_MIN//2
    for i in range(firs+1, len(dm)-firs-1):
        if dm.loc[i, 'match'] == 'crimson' or dm.loc[i, 'match'] == 'yellow':
            flag = True
            for j in range(i-firs, i+firs):
                if j != i:
                    if dm.loc[j, 'match'] != 'limegreen':
                        flag = False
                        break
            if flag:
                dm.loc[i, 'match'] = 'limegreen'
                
    dc = dm[dm['match'] == 'crimson']
    dc = dc.reset_index(drop=True)
    
    if len(dc) > 0:
        if len(dc) == 1:
            dm.loc[(dm['position']==dc.loc[0,'position']),'match']='yellow'
        else:
            mm_dst = MM_DIST * 1000
        
            for i in range(len(dc)):
                if i == 0:
                    if dc.loc[1,'position'] - dc.loc[0,'position'] > mm_dst:
                        dm.loc[(dm['position']==dc.loc[0,'position']),'match']='yellow'
                elif i == len(dc)-1:
                    if dc.loc[i,'position'] - dc.loc[i-1,'position'] > mm_dst:
                        dm.loc[(dm['position']==dc.loc[i,'position']),'match']='yellow'
                elif dc.loc[i,'position'] - dc.loc[i-1,'position'] > mm_dst and \
                    dc.loc[i+1,'position'] - dc.loc[i,'position'] > mm_dst:
                        dm.loc[(dm['position']==dc.loc[i,'position']),'match']='yellow'

    return dm
   
if __name__ == "__main__":
    start_time = time.time()    
    FILES_PATH = os.path.normpath(FILES_PATH)
    MAP_PATH = os.path.normpath(MAP_PATH)

    if len(CHROMOSOMES) > 0 and (CHROMOSOMES[-1] > 23 or \
                0 in CHROMOSOMES) and COUSINS == []:            
        print("\nChromosome Number must be 1 to 23. Try again.")
        sys.exit()

    if len(SIBLINGS) < 2 and len(PHASED_FILES) == 0 and COUSINS == []:        
            print("\nThere must be at least two SIBLINGS. Try again.")
            sys.exit()
            
    if len(PHASED_FILES) < 2 and len(SIBLINGS) == 0 and COUSINS == []:
            print("\nThere must be at least two PHASED_FILES. Try again.")
            sys.exit()
            
    print("\nPlease wait while DNA files are loaded...\n")
    
    # Initialize variables
    xlfilnme = EXCEL_FILE_NAME
    wdir = os.path.normpath(WORKING_DIRECTORY) + "/"
    
    # Get Excel file name.
    xlpath = wdir + xlfilnme
    xlname = xlpath + ".xlsx"
    
    if LINEAR_CHROMOSOME and COUSINS == []:
        AUTO_REC_PNTS = False
    
    if COUSINS == []:           
        # Create an new Excel file and delete 'Sheet'.
        wb = Workbook()
        ws = wb.active
        del wb["Sheet"]        
    else:
        try:
            wb = load_workbook(xlname)
        except:
            print(f'File {xlname} does not exist. Enter an existing file.')
            sys.exit()
            
    # Load Minimal Map.
    dmap_source = load_map()
    
    # Chromosome lengths (Build 37). 
    chr_lens = [249250621, 243199373, 198022430, 191154276, 180915260, 171115067,
     159138663, 146364022, 141213431, 135534747, 135006516, 133851895,
     115169878, 107349540, 102531392, 90354753, 81195210, 78077248, 59128983,
     63025520, 48129895, 51304566, 155270560]
    
    # Determine match pairs
    if COUSINS  !=  []:
            
        chroms = wb.sheetnames
        CHROMOSOMES = []
        SIBLINGS = []
        
        ws = wb[chroms[0]]
        
        # Retrieve mode.
        if ws.cell(4,1).value == 1:
            LINEAR_CHROMOSOME = False
            CHROM_TRUE_SIZE = False
        elif ws.cell(4,1).value == 2:
            LINEAR_CHROMOSOME = False
            CHROM_TRUE_SIZE = True
        elif ws.cell(4,1).value == 3:
            LINEAR_CHROMOSOME = True
            AUTO_REC_PNTS = False
            
        # Retrieve variables.
        ARP_TOLERANCE = ws.cell(2,1).value
        RESOLUTION = ws.cell(1,1).value
        
        if ws.cell(5,1).value == 1:
            AUTO_REC_PNTS = True
        else:
            AUTO_REC_PNTS = False
        
        for i in range(1,1000):
            if ws.cell(i,7).value != None and ws.cell(i,7).value != 'Column Width':
                SIBLINGS.append(ws.cell(i,7).value)
                
            for name in SIBLINGS:
                if '-' in name:
                    SIBLINGS.remove(name)
                        
        if SIBLINGS == []:
            print('\nThere are no siblings to compare to.')
            sys.exit()
                        
        match_pairs = []
       
        for ind in SIBLINGS:
            for cousin in COUSINS:
                if ind != cousin:
                    comb = (ind,cousin)
                    match_pairs.append(comb)
                    
        for chrom in chroms:
            CHROMOSOMES.append(int(chrom[3:]))
        
    else: 
        
        match_pairs = list(combinations(SIBLINGS, 2))
                                    
        if PHASED_FILES != []:                                 
            phased_pairs = list(combinations(PHASED_FILES, 2))
            match_pairs = match_pairs + phased_pairs
            
        if EVIL_TWINS != []:
            for sib in SIBLINGS:
                for etw in EVIL_TWINS:
                    if sib not in etw:
                        comb = (sib,etw)
                        match_pairs.append(comb)
            
    for chrom in range(1, 24):
        # Initialize recombination point list.
        rps_list = []
        
        # Initialize rp assigment list.
        rnames_list = []

        if CHROMOSOMES == [] or chrom in CHROMOSOMES:
            dmap = dmap_source[dmap_source["Chromosome"] == chrom]
            dmap = dmap.reset_index(drop=True)
            
            if COUSINS  !=  []:
                ws = wb[f'Chr{chrom}']
                
                im_width = ws._images[1].width
                
                last_col = ws.cell(3,1).value
                                                        
            else:
                ws = wb.create_sheet(f"Chr{chrom}")
                ws.cell(1,1).value = RESOLUTION
                ws.cell(2,1).value = ARP_TOLERANCE

                # Save mode.
                if not LINEAR_CHROMOSOME and not CHROM_TRUE_SIZE:
                    ws.cell(4,1).value = 1
                elif not LINEAR_CHROMOSOME and CHROM_TRUE_SIZE:
                    ws.cell(4,1).value = 2
                elif LINEAR_CHROMOSOME:
                    ws.cell(4,1).value = 3
                    
                if AUTO_REC_PNTS:
                    ws.cell(5,1).value = 1
                                    
            col = cs(FREEZE_COLUMN) 
            col_to_freeze = cl(col+1)   
            ws.freeze_panes = ws[col_to_freeze + '1']
            
            format_sheet(ws)
    
            # Compare match pairs. Return Tables and images of results.
            
            dtot = pd.DataFrame()
            dxtot = pd.DataFrame()
            dstot = pd.DataFrame()
            comparisons = 0
            current_time = time.time()
            
            for pair in match_pairs:
                
                initial_time = time.time()
                
                dm = load_dna_files(pair, chrom)
                                
                pair_name = pair[0] + "-" + pair[1]
    
                if len(dm) == 0:
                    print("\nDNA file not found. Check DNA files folder and file formats.")
                    sys.exit()
    
                dx, ds = scan_genomes(dm, chrom)                                    
                
                paste_tables(dx, ds, pair_name)
                
                dx['pair'] = pair_name
                ds['pair'] = pair_name
                
                if len(dxtot) == 0:
                    dxtot = dx
                else:
                    dxtot = pd.concat([dxtot, dx], ignore_index=True)
                    
                if len(dstot) == 0:
                    dstot = ds
                else:
                    dstot = pd.concat([dstot, ds], ignore_index=True)

                cols = list(dm.columns)           
                cols.remove('position')
                cols.remove('match')
                
                dm = dm.drop(cols, axis=1)
                
                dm = dm.rename(columns={'match':pair_name})
                
                if len(dtot) == 0:
                    dtot = dm
                else:
                    if MERGE_FILES:
                        dtot = pd.merge(dtot, dm, on=('position')) 
                    else:
                        dtot = pd.merge(dtot, dm, on=('position'), how='outer')
                
                dtot = dtot.fillna('grey')        

                comparison_time = time.time() - initial_time
                
                comparisons+=1
                
                if SHOW_MATCH_PAIR_PROGRESS:
                    print(f'{pair_name} tables Chromosome {chrom} complete.')
                  
                    if SHOW_TIMES:
                        print(f'Comparisons = {comparisons}, elapsed time = {comparison_time:.1f} seconds') 

            comparisons = 0 
            
            finish_time = time.time() - current_time
            
            if not SHOW_MATCH_PAIR_PROGRESS:            
                print(f'Chromosome {chrom} tables complete')
                if SHOW_TIMES:
                    print(f'Elapsed time = {finish_time:.1f} seconds') 

            
            for q,pair in enumerate(match_pairs):
                    
                initial_time = time.time()  
                
                pair_name = pair[0] + "-" + pair[1]
                
                dplot, rps_list, rnames_list = get_dplot(q, dtot, dxtot, dstot, pair_name)
                
                if q == 0:
                    if SCALE_ON:
                        get_scale(dplot, chrom)
                                                    
                get_image(dplot, pair_name, chrom)
                                                
                paste_image(pair_name, chrom, q, dxtot)
                
                comparison_time = time.time() - initial_time 
                
                comparisons+=1
                
                if SHOW_MATCH_PAIR_PROGRESS:
                    print(f'{pair_name} Chromosome {chrom} complete.')
                  
                    if SHOW_TIMES:
                        print(f'Comparisons = {comparisons}, elapsed time = {comparison_time:.1f} seconds.') 
            
            if COUSINS == []:                      
                if AUTO_REC_PNTS:                
                    last_col = do_arp(rps_list, rnames_list)
                    ws.cell(3,1).value = last_col
                    add_chroms(ws, last_col) 
                    add_borders(ws, last_col)
                else:
                    add_chroms(ws, 50)
                    add_borders(ws, 50)
            else:
                if AUTO_REC_PNTS: 
                    add_borders(ws, last_col)
                else:
                    add_borders(ws, 50)
                               
            finish_time = time.time() - current_time
            
            print(f'\nChromosome {chrom} complete.\n') 
            
            if SHOW_TIMES:
                print(f'Elapsed time = {finish_time:.1f} seconds.\n') 
                        
    wb.save(xlname)        
                
    # Delete temporary images
    delete_images()
    
    total_time = time.time() - start_time
    
    print(f'\nTotal elapsed time = {total_time//60:.0f} minutes {total_time % 60: .0f} seconds')
                
    print('\nFinished')



